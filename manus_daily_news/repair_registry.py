#!/usr/bin/env python3
"""One-time repair of the 35 column-shifted rows in the Jul-30 source registry.

Corruption pattern: `robotics_relevance` prose containing commas was split
across robotics_relevance / priority_tier / collection_route, shifting the
real priority (P0/P1/P2), collection route, and watch keywords rightward.
The fragments are all present, so the row is reconstructed losslessly by
rejoining the fragments and reassigning the displaced values.

Input : RAI_FirstHand_Source_Registry_500_20260730.xlsx (sheet "All Sources")
Output: registry_v4_20260805.csv (canonical)
        RAI_FirstHand_Source_Registry_v4_20260805.xlsx (Summary + All Sources)

Run from scripts/manus_daily_news/. Requires openpyxl.
"""
import csv
import re
import sys
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter

SRC_XLSX = "RAI_FirstHand_Source_Registry_500_20260730.xlsx"
OUT_CSV = "registry_v4_20260805.csv"
OUT_XLSX = "RAI_FirstHand_Source_Registry_v4_20260805.xlsx"

CSV_HEADER = [
    "region", "source_name", "source_name_local", "type", "subtype",
    "url_news", "url_home", "language", "rss_or_api", "update_frequency",
    "robotics_relevance", "priority_tier", "collection_route",
    "watch_keywords", "verification_role", "ticker", "origin", "notes",
]

TIERS = {"P0", "P1", "P2"}
ROUTE_RE = re.compile(
    r"^(page scrape|RSS|API|WeChat|search query)([ +].*)?$", re.IGNORECASE
)
ORIGIN_RE = re.compile(r"Registry v\d|Expansion|Jul \d{2}|Aug \d{2}")


def is_route(v):
    return bool(v) and bool(ROUTE_RE.match(v.strip()))


def repair_row(vals):
    """vals: list of 18 cell values. Returns (fixed_vals, was_bad)."""
    vals = [(v if v is not None else "") for v in vals]
    vals = [str(v).strip() for v in vals]
    if vals[11] in TIERS:
        return vals, False

    tail = vals[10:]
    tier_idx = next((i for i, v in enumerate(tail) if v in TIERS), None)
    if tier_idx is None:
        return vals, None  # unrepairable, needs manual review

    relevance = ", ".join(v for v in tail[:tier_idx] if v)
    priority = tail[tier_idx]
    rest = [v for v in tail[tier_idx + 1:] if v]

    route = next((v for v in rest if is_route(v)), "")
    rest = [v for v in rest if v != route]
    origin = next((v for v in rest if ORIGIN_RE.search(v)), "")
    rest = [v for v in rest if v != origin]
    watch = rest[0] if rest else ""
    notes = rest[1] if len(rest) > 1 else ""

    fixed = vals[:10] + [relevance, priority, route, watch, "", "", origin, notes]
    return fixed, True


# field-level fixes for rows that were entered inconsistently (not shifted):
# source_name -> {column_index: new_value}
NORMALIZE = {
    "National Local Joint Humanoid Robot Innovation Center": {
        5: "https://mp.weixin.qq.com/",
        17: "WeChat search: 国地共建人形机器人创新中心",
    },
    "Oceaneering International": {12: "page scrape + IR calendar"},
    "Richtech Robotics": {12: "page scrape + exchange filings"},
    "Nasdaq Nordic": {12: "RSS + page scrape"},
    "Star Plus Legend Holdings Limited": {12: "page scrape + HKEX filings"},
    "ASEAN Secretariat": {12: "page scrape"},
}


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, read_only=True)
    ws = wb["All Sources"]
    rows = list(ws.values)

    fixed_rows, repaired, manual = [], [], []
    for r in rows[1:]:
        fixed, was_bad = repair_row(list(r))
        if was_bad is None:
            manual.append(fixed)
        elif was_bad:
            repaired.append(fixed[1])
        for col, val in NORMALIZE.get(fixed[1], {}).items():
            if col == 17 and fixed[col]:
                fixed[col] = f"{fixed[col]}; {val}"
            else:
                fixed[col] = val
        fixed_rows.append(fixed)

    print(f"rows: {len(fixed_rows)}, repaired: {len(repaired)}, manual: {len(manual)}")
    if repaired:
        print("repaired:", ", ".join(repaired))
    if manual:
        print("NEEDS MANUAL REVIEW:")
        for m in manual:
            print("  ", m[:3])
        sys.exit(1)

    tiers = Counter(r[11] for r in fixed_rows)
    assert set(tiers) <= TIERS, f"unexpected tiers remain: {tiers}"

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADER)
        w.writerows(fixed_rows)
    print(f"wrote {OUT_CSV}  tiers={dict(tiers)}")

    out = openpyxl.Workbook()
    summ = out.active
    summ.title = "Summary"
    summ.append(["Registry v4 — repaired 2026-08-05 (35 column-shifted rows fixed)"])
    summ.append([])
    summ.append(["Priority", "Count"])
    for t in sorted(tiers):
        summ.append([t, tiers[t]])
    summ.append([])
    summ.append(["Region", "Count"])
    for reg, n in Counter(r[0] for r in fixed_rows).most_common():
        summ.append([reg, n])

    all_ws = out.create_sheet("All Sources")
    all_ws.append(CSV_HEADER)
    for r in fixed_rows:
        all_ws.append(r)
    for i in range(1, len(CSV_HEADER) + 1):
        all_ws.column_dimensions[get_column_letter(i)].width = 22
    out.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
