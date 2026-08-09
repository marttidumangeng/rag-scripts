import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REG_DIR = '/home/ubuntu/rai_registry/regions'
ORDER = ['China', 'Japan', 'South_Korea', 'United_States', 'Europe', 'Taiwan', 'Hong_Kong', 'International']
FIELDS = ['region','source_name','source_name_local','type','subtype','url_news','url_home','language',
          'rss_or_api','update_frequency','robotics_relevance','priority_tier','collection_route',
          'watch_keywords','verification_role','notes']

data = {}
for reg in ORDER:
    path = f'{REG_DIR}/{reg}_sources_csv.csv'
    with open(path, newline='', encoding='utf-8-sig') as f:
        rows = []
        for row in csv.DictReader(f):
            row.pop(None, None)
            rows.append({k: (row.get(k) or '').strip() for k in FIELDS})
        data[reg] = rows

wb = Workbook()

hdr_fill = PatternFill('solid', fgColor='1F3864')
hdr_font = Font(color='FFFFFF', bold=True, size=10)
p0_fill = PatternFill('solid', fgColor='FCE4D6')
wire_fill = PatternFill('solid', fgColor='E2EFDA')
thin = Border(*[Side(style='thin', color='D9D9D9')]*4)

# Summary sheet
ws = wb.active
ws.title = 'Summary'
ws['A1'] = 'RAI First-Hand Primary Source Registry — Master (July 28, 2026)'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'Definition: every counted source ORIGINATES information (government, disclosure, patents, procurement, statistics, associations, standards, companies, research institutes, exhibitions, preprints). Wire services are included only as clearly-labeled verification-tier entries and are not counted toward the 30-per-region minimum.'
ws['A2'].alignment = Alignment(wrap_text=True)
ws.merge_cells('A2:F2')
ws.row_dimensions[2].height = 45

headers = ['Region', 'First-hand sources', 'Verification wires', 'Total rows', 'P0 daily-scan', 'Top origination sources']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = thin

TOP_ORIG = {
 'China': 'MIIT press releases; CNINFO/SSE-SZSE filings; company newsrooms (Unitree/UBTECH/AgiBot)',
 'Japan': 'TDnet timely disclosure; EDINET; METI press releases',
 'South_Korea': 'DART filings; MOTIE press releases; KAIST news',
 'United_States': 'SEC EDGAR full-text search; FDA 510(k) database; arXiv cs.RO',
 'Europe': 'European Commission digital-strategy news; Fraunhofer/DLR releases; ABB/KUKA newsrooms',
 'Taiwan': 'TWSE MOPS filings; NSTC announcements; ITRI news',
 'Hong_Kong': 'HKEXnews filings; ITC press releases; HKSTP news',
 'International': 'IFR press releases; arXiv cs.RO; ISO TC 299 standards tracker',
}

r = 5
tot_fh = tot_w = 0
for reg in ORDER:
    rows = data[reg]
    wires = sum(1 for x in rows if 'wire' in x['type'].lower())
    fh = len(rows) - wires
    p0 = sum(1 for x in rows if x['priority_tier'].strip().upper().startswith('P0'))
    tot_fh += fh; tot_w += wires
    vals = [reg.replace('_', ' '), fh, wires, len(rows), p0, TOP_ORIG[reg]]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = thin
    r += 1
for c, v in enumerate(['TOTAL', tot_fh, tot_w, tot_fh + tot_w, '', ''], 1):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True); cell.border = thin

widths = [16, 18, 16, 12, 14, 70]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Legend
ws.cell(row=r+2, column=1, value='Legend: orange rows = P0 daily-scan priority; green rows = verification-only wires (not first-hand).').font = Font(italic=True, size=9)
ws.cell(row=r+3, column=1, value='Priority: P0 = scan daily, P1 = scan 2-3x/week, P2 = event-driven/monthly.').font = Font(italic=True, size=9)

# Region sheets
COL_W = {'region':10,'source_name':32,'source_name_local':22,'type':13,'subtype':22,'url_news':50,'url_home':32,
         'language':9,'rss_or_api':30,'update_frequency':13,'robotics_relevance':50,'priority_tier':9,
         'collection_route':16,'watch_keywords':28,'verification_role':28,'notes':45}

for reg in ORDER:
    ws = wb.create_sheet(reg.replace('_', ' ')[:31])
    for c, h in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = thin
        ws.column_dimensions[get_column_letter(c)].width = COL_W[h]
    # sort: type groups then priority
    type_order = {'Government':0,'Disclosure':1,'Patent':2,'Procurement':3,'Statistics':4,'Association':5,
                  'Standards':6,'Research':7,'University':8,'Company':9,'Exhibition':10,'Preprint':11}
    def key(x):
        t = x['type'].split('(')[0].strip()
        return (type_order.get(t, 12), x['priority_tier'], x['source_name'].lower())
    rows = sorted(data[reg], key=key)
    for r0, row in enumerate(rows, 2):
        is_wire = 'wire' in row['type'].lower()
        is_p0 = row['priority_tier'].strip().upper().startswith('P0')
        for c, h in enumerate(FIELDS, 1):
            cell = ws.cell(row=r0, column=c, value=row[h])
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(size=9)
            if is_wire:
                cell.fill = wire_fill
            elif is_p0:
                cell.fill = p0_fill
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(FIELDS))}{len(rows)+1}'

out_xlsx = '/home/ubuntu/rai_registry/RAI_FirstHand_Source_Registry_Master_20260728.xlsx'
wb.save(out_xlsx)
print('saved', out_xlsx)

# Also flat CSV
out_csv = '/home/ubuntu/rai_registry/tier1_firsthand_source_registry_20260728.csv'
with open(out_csv, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=FIELDS)
    w.writeheader()
    for reg in ORDER:
        for row in data[reg]:
            w.writerow(row)
print('saved', out_csv)
for reg in ORDER:
    rows = data[reg]
    wires = sum(1 for x in rows if 'wire' in x['type'].lower())
    print(reg, 'first-hand:', len(rows)-wires, 'wires:', wires)
