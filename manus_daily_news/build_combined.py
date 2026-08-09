#!/usr/bin/env python3
"""Merge 124 newly verified sources with the existing 317-source registry
into a combined master CSV + formatted Excel workbook."""
import csv, json
from collections import Counter, OrderedDict

BASE = '/home/ubuntu/rai_registry'
EXISTING = f'{BASE}/tier1_firsthand_source_registry_20260728.csv'
NEW_JSON = f'{BASE}/verify_new_firsthand_sources.json'
OUT_CSV = f'{BASE}/tier1_firsthand_source_registry_combined_20260730.csv'
OUT_XLSX = f'{BASE}/RAI_FirstHand_Source_Registry_Combined_20260730.xlsx'

COLS = ['region','source_name','source_name_local','type','subtype','url_news','url_home',
        'language','rss_or_api','update_frequency','robotics_relevance','priority_tier',
        'collection_route','watch_keywords','verification_role','ticker','origin','notes']

# ---- load existing ----
rows = []
with open(EXISTING, newline='', encoding='utf-8') as f:
    for r in csv.DictReader(f):
        r = dict(r)
        r['ticker'] = ''
        r['origin'] = 'Registry v1 (Jul 28)'
        rows.append({c: r.get(c, '') for c in COLS})

existing_names = {r['source_name'].strip().lower() for r in rows}
existing_homes = {r['url_home'].strip().rstrip('/').lower() for r in rows if r['url_home']}

# ---- load new ----
data = json.load(open(NEW_JSON))['results']
new_rows, skipped = [], []
for item in data:
    o = item['output']
    name = o['source_name'].strip()
    home = (o.get('url_home') or '').strip().rstrip('/').lower()
    if name.lower() in existing_names or (home and home in existing_homes):
        skipped.append(name); continue
    # normalize type into type/subtype
    t = o['type']
    if t == 'Listed Company':
        typ, sub = 'Company', 'Listed company IR/newsroom'
    elif t == 'Company':
        typ, sub = 'Company', 'Company newsroom'
    elif t in ('Government', 'Government Agency', 'Government Programme'):
        typ, sub = 'Government', t if t != 'Government' else 'Government press releases'
    elif t == 'Regulator':
        typ, sub = 'Government', 'Regulator'
    elif t == 'Research Institute':
        typ, sub = 'Research', 'Research institute'
    elif t == 'Exhibition':
        typ, sub = 'Exhibition', 'Exhibition press office'
    elif t == 'Incubator':
        typ, sub = 'Ecosystem', 'Incubator/cluster'
    else:
        typ, sub = t, ''
    new_rows.append({
        'region': o['region'],
        'source_name': name,
        'source_name_local': '' if o.get('source_name_local','NA') in ('NA','') else o['source_name_local'],
        'type': typ,
        'subtype': sub,
        'url_news': o['url_news'],
        'url_home': o['url_home'],
        'language': o['language'],
        'rss_or_api': o['rss_or_api'],
        'update_frequency': o['update_frequency'],
        'robotics_relevance': o['robotics_relevance'],
        'priority_tier': o['priority_tier'],
        'collection_route': o['collection_route'],
        'watch_keywords': o['watch_keywords'],
        'verification_role': 'primary announcement',
        'ticker': '' if o.get('ticker','NA') in ('NA','') else o['ticker'],
        'origin': 'Site review v2 (Jul 30)',
        'notes': '' if o.get('notes','None') in ('None','') else o['notes'],
    })
    existing_names.add(name.lower())
    if home: existing_homes.add(home)

print(f'existing: {len(rows)}, new added: {len(new_rows)}, skipped dupes: {len(skipped)}')
if skipped: print('skipped:', skipped)

all_rows = rows + new_rows
REGION_ORDER = ['China','Japan','South Korea','United States','Europe','Taiwan','Hong Kong','International']
PRIO = {'P0':0,'P1':1,'P2':2}
all_rows.sort(key=lambda r: (REGION_ORDER.index(r['region']) if r['region'] in REGION_ORDER else 99,
                             PRIO.get(r['priority_tier'], 3), r['type'], r['source_name']))

with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=COLS)
    w.writeheader(); w.writerows(all_rows)
print('wrote', OUT_CSV, len(all_rows), 'rows')

# ---- Excel ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
P0_FILL = PatternFill('solid', fgColor='FFF2CC')
NEW_FONT = Font(color='C00000', size=9)
THIN = Border(*[Side(style='thin', color='D9D9D9')]*4)

HEADERS = ['Region','Source Name','Local Name','Type','Subtype','News URL','Homepage','Lang',
           'RSS/API','Frequency','Robotics Relevance','Priority','Collection Route',
           'Watch Keywords','Verification Role','Ticker','Origin','Notes']
KEYS = COLS
WIDTHS = [11,30,20,12,22,45,32,8,22,10,45,8,20,28,20,14,18,28]

# Summary sheet
ws = wb.active; ws.title = 'Summary'
ws['A1'] = 'RAI First-Hand Source Registry — Combined Master (v2, Jul 30 2026)'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'Registry v1 (Jul 28): 324 sources | Site-review additions (Jul 30): %d | Total: %d' % (len(new_rows), len(all_rows))
ws['A2'].font = Font(size=10, italic=True)
hdr = ['Region','Total','P0','P1','P2','Government','Company','Research','Association','Exhibition','Other','New (v2)']
for j,h in enumerate(hdr,1):
    c = ws.cell(4, j, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
r_i = 5
for reg in REGION_ORDER:
    rs = [r for r in all_rows if r['region']==reg]
    tc = Counter(r['type'] for r in rs)
    pc = Counter(r['priority_tier'] for r in rs)
    other = len(rs) - tc.get('Government',0)-tc.get('Company',0)-tc.get('Research',0)-tc.get('Association',0)-tc.get('Exhibition',0)
    newc = sum(1 for r in rs if r['origin'].startswith('Site review'))
    vals = [reg, len(rs), pc.get('P0',0), pc.get('P1',0), pc.get('P2',0),
            tc.get('Government',0), tc.get('Company',0), tc.get('Research',0),
            tc.get('Association',0), tc.get('Exhibition',0), other, newc]
    for j,v in enumerate(vals,1):
        c = ws.cell(r_i, j, v); c.border = THIN
    r_i += 1
tot = ['TOTAL', len(all_rows),
       sum(1 for r in all_rows if r['priority_tier']=='P0'),
       sum(1 for r in all_rows if r['priority_tier']=='P1'),
       sum(1 for r in all_rows if r['priority_tier']=='P2'),
       sum(1 for r in all_rows if r['type']=='Government'),
       sum(1 for r in all_rows if r['type']=='Company'),
       sum(1 for r in all_rows if r['type']=='Research'),
       sum(1 for r in all_rows if r['type']=='Association'),
       sum(1 for r in all_rows if r['type']=='Exhibition'),
       0, len(new_rows)]
tot[10] = tot[1]-sum(tot[5:10])
for j,v in enumerate(tot,1):
    c = ws.cell(r_i, j, v); c.font = Font(bold=True); c.border = THIN
for j,w_ in enumerate([14,8,6,6,6,12,10,10,12,11,8,9],1):
    ws.column_dimensions[get_column_letter(j)].width = w_
ws.cell(r_i+2,1,'Legend: yellow rows = P0 daily-scan sources; red text in Origin = new additions from the robotaigeek.com site review.').font = Font(size=9, italic=True)

# All Sources sheet + per-region sheets
def fill_sheet(ws, rows_):
    for j,h in enumerate(HEADERS,1):
        c = ws.cell(1, j, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical='center')
    for i,r in enumerate(rows_, 2):
        is_new = r['origin'].startswith('Site review')
        for j,k in enumerate(KEYS,1):
            c = ws.cell(i, j, r[k]); c.border = THIN
            c.font = Font(size=9) if not (is_new and k=='origin') else NEW_FONT
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if r['priority_tier']=='P0': c.fill = P0_FILL
    for j,w_ in enumerate(WIDTHS,1):
        ws.column_dimensions[get_column_letter(j)].width = w_
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{len(rows_)+1}'

fill_sheet(wb.create_sheet('All Sources'), all_rows)
for reg in REGION_ORDER:
    nm = reg if len(reg) <= 28 else reg[:28]
    fill_sheet(wb.create_sheet(nm), [r for r in all_rows if r['region']==reg])

wb.save(OUT_XLSX)
print('wrote', OUT_XLSX)
print(Counter(r['region'] for r in all_rows))
