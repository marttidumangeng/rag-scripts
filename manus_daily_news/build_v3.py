#!/usr/bin/env python3
"""Merge batch-3 (exchanges, SEA gov, associations, authorities) into the
combined registry -> v3 master CSV + Excel targeting 500+ sources."""
import csv, json
from collections import Counter

BASE = '/home/ubuntu/rai_registry'
EXISTING = f'{BASE}/tier1_firsthand_source_registry_combined_20260730.csv'
NEW_JSON = '/home/ubuntu/verify_exchange_sea_authority_sources.json'
OUT_CSV = f'{BASE}/tier1_firsthand_source_registry_v3_20260730.csv'
OUT_XLSX = f'{BASE}/RAI_FirstHand_Source_Registry_500_20260730.xlsx'

COLS = ['region','source_name','source_name_local','type','subtype','url_news','url_home',
        'language','rss_or_api','update_frequency','robotics_relevance','priority_tier',
        'collection_route','watch_keywords','verification_role','ticker','origin','notes']

rows = []
with open(EXISTING, newline='', encoding='utf-8') as f:
    for r in csv.DictReader(f):
        rows.append({c: r.get(c, '') for c in COLS})

existing_names = {r['source_name'].strip().lower() for r in rows}
existing_news = {r['url_news'].strip().rstrip('/').lower() for r in rows if r['url_news']}

data = json.load(open(NEW_JSON))['results']
new_rows, excluded, dupes = [], [], []
for item in data:
    o = item['output']
    name = o['source_name'].strip()
    if o.get('exclude'):
        excluded.append((name, o.get('notes','')))
        continue
    nu = (o.get('url_news') or '').strip().rstrip('/').lower()
    if name.lower() in existing_names or (nu and nu in existing_news):
        dupes.append(name); continue
    t = o['type']
    if t == 'Exchange':
        typ, sub = 'Exchange', 'Exchange newsroom / listings'
    elif t == 'Government':
        typ, sub = 'Government', 'Government press releases'
    elif t == 'Association':
        typ, sub = 'Association', 'Industry association'
    elif t == 'Research Institute':
        typ, sub = 'Research', 'Research institute'
    elif t == 'Standards Body':
        typ, sub = 'Standards', 'Standards body'
    elif t == 'Independent Authority':
        typ, sub = 'Authority', 'Independent authority'
    else:
        typ, sub = t, ''
    new_rows.append({
        'region': o['region'],
        'source_name': name,
        'source_name_local': '' if o.get('source_name_local','NA') in ('NA','') else o['source_name_local'],
        'type': typ, 'subtype': sub,
        'url_news': o['url_news'], 'url_home': o['url_home'],
        'language': o['language'], 'rss_or_api': o['rss_or_api'],
        'update_frequency': o['update_frequency'],
        'robotics_relevance': o['robotics_relevance'],
        'priority_tier': o['priority_tier'],
        'collection_route': o['collection_route'],
        'watch_keywords': o['watch_keywords'],
        'verification_role': 'primary disclosure/announcement',
        'ticker': '',
        'origin': 'Exchange/SEA batch v3 (Jul 30)',
        'notes': '' if o.get('notes','None') in ('None','') else o['notes'],
    })
    existing_names.add(name.lower())
    if nu: existing_news.add(nu)

print(f'existing: {len(rows)}, new: {len(new_rows)}, excluded: {len(excluded)}, dupes: {len(dupes)}')
for n, why in excluded: print('  EXCLUDED:', n, '|', why[:90])
for n in dupes: print('  DUPE:', n)

all_rows = rows + new_rows
total = len(all_rows)
print('TOTAL:', total)

REGION_ORDER = ['China','Japan','South Korea','United States','Europe','Taiwan','Hong Kong','International']
PRIO = {'P0':0,'P1':1,'P2':2}
all_rows.sort(key=lambda r: (REGION_ORDER.index(r['region']) if r['region'] in REGION_ORDER else 99,
                             PRIO.get(r['priority_tier'], 3), r['type'], r['source_name']))

with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=COLS)
    w.writeheader(); w.writerows(all_rows)
print('wrote', OUT_CSV)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
P0_FILL = PatternFill('solid', fgColor='FFF2CC')
V3_FONT = Font(color='006100', size=9, bold=True)
V2_FONT = Font(color='C00000', size=9)
THIN = Border(*[Side(style='thin', color='D9D9D9')]*4)

HEADERS = ['Region','Source Name','Local Name','Type','Subtype','News URL','Homepage','Lang',
           'RSS/API','Frequency','Robotics Relevance','Priority','Collection Route',
           'Watch Keywords','Verification Role','Ticker','Origin','Notes']
WIDTHS = [11,30,20,12,24,45,32,8,22,10,45,8,20,28,22,14,22,28]

ws = wb.active; ws.title = 'Summary'
ws['A1'] = 'RAI First-Hand Source Registry — 500 Master (v3, Jul 30 2026)'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = f'v1 registry (Jul 28): 324 | Site-review additions (v2): 124 | Exchange/SEA/Authority batch (v3): {len(new_rows)} | TOTAL: {total}'
ws['A2'].font = Font(size=10, italic=True)
hdr = ['Region','Total','P0','P1','P2','Government','Company','Exchange','Research','Association','Standards/Authority','Other','v3 New']
for j,h in enumerate(hdr,1):
    c = ws.cell(4, j, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
r_i = 5
def cnt(rs):
    tc = Counter(r['type'] for r in rs)
    pc = Counter(r['priority_tier'] for r in rs)
    sa = tc.get('Standards',0)+tc.get('Authority',0)
    known = tc.get('Government',0)+tc.get('Company',0)+tc.get('Exchange',0)+tc.get('Research',0)+tc.get('Association',0)+sa
    return [len(rs), pc.get('P0',0), pc.get('P1',0), pc.get('P2',0),
            tc.get('Government',0), tc.get('Company',0), tc.get('Exchange',0),
            tc.get('Research',0), tc.get('Association',0), sa, len(rs)-known,
            sum(1 for r in rs if r['origin'].startswith('Exchange/SEA'))]
for reg in REGION_ORDER:
    rs = [r for r in all_rows if r['region']==reg]
    for j,v in enumerate([reg]+cnt(rs),1):
        c = ws.cell(r_i, j, v); c.border = THIN
    r_i += 1
for j,v in enumerate(['TOTAL']+cnt(all_rows),1):
    c = ws.cell(r_i, j, v); c.font = Font(bold=True); c.border = THIN
for j,w_ in enumerate([14,7,6,6,6,12,10,10,10,12,18,8,8],1):
    ws.column_dimensions[get_column_letter(j)].width = w_
ws.cell(r_i+2,1,'Legend: yellow rows = P0 daily-scan. Origin column: black = v1 (Jul 28), red = v2 site-review, green bold = v3 exchange/SEA/authority batch.').font = Font(size=9, italic=True)

def fill_sheet(ws, rows_):
    for j,h in enumerate(HEADERS,1):
        c = ws.cell(1, j, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical='center')
    for i,r in enumerate(rows_, 2):
        for j,k in enumerate(COLS,1):
            c = ws.cell(i, j, r[k]); c.border = THIN
            if k=='origin' and r[k].startswith('Exchange/SEA'): c.font = V3_FONT
            elif k=='origin' and r[k].startswith('Site review'): c.font = V2_FONT
            else: c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if r['priority_tier']=='P0': c.fill = P0_FILL
    for j,w_ in enumerate(WIDTHS,1):
        ws.column_dimensions[get_column_letter(j)].width = w_
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{len(rows_)+1}'

fill_sheet(wb.create_sheet('All Sources'), all_rows)
for reg in REGION_ORDER:
    fill_sheet(wb.create_sheet(reg), [r for r in all_rows if r['region']==reg])
# dedicated exchange view
fill_sheet(wb.create_sheet('Exchanges & Disclosure'), [r for r in all_rows if r['type']=='Exchange' or 'disclosure' in (r['subtype'] or '').lower() or r['source_name'] in ('SEC EDGAR','CNINFO','HKEXnews','TDnet','EDINET','Financial Supervisory Service (FSS) DART','Korea Exchange (KRX) KIND','TWSE Market Observation Post System (MOPS)','SSE','SZSE','BSE')])

wb.save(OUT_XLSX)
print('wrote', OUT_XLSX)
print(Counter(r['region'] for r in all_rows))
