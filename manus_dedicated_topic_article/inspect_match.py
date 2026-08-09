from openpyxl import load_workbook
from datetime import datetime, date
path = '/home/ubuntu/projects/dedicated-topic-article-090d06d7/ARPI_Content_Calendar_Strategy_v6.xlsx'
wb = load_workbook(path, data_only=False)
for s in wb.sheetnames:
    ws = wb[s]
    print('\nSHEET', s)
    if s == '4. Week-by-Week (Wk 1–8)':
        for r in range(45, 53):
            print('ROW', r, [(ws.cell(r,c).coordinate, ws.cell(r,c).value, ws.cell(r,c).number_format, ws.cell(r,c).style_id) for c in range(1, ws.max_column+1) if ws.cell(r,c).value is not None])
        print('MERGED', list(ws.merged_cells.ranges))
    if s == '5. Content Framework':
        for r in range(70, 83):
            print('ROW', r, [ws.cell(r,c).value for c in range(1, ws.max_column+1)])
        print('MERGED', list(ws.merged_cells.ranges))
