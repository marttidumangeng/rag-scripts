from datetime import datetime, date
from openpyxl import load_workbook

path = '/home/ubuntu/projects/dedicated-topic-article-090d06d7/ARPI_Content_Calendar_Strategy_v6.xlsx'
run_date = date(2026, 8, 6)
wb = load_workbook(path, data_only=False)
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'SHEET {name!r} rows={ws.max_row} cols={ws.max_column}')

ws = wb['4. Week-by-Week (Wk 1–8)']
print('TAB4_HEADERS_AND_MATCHES')
for r in ws.iter_rows(values_only=True):
    vals = list(r)
    if any(v is not None for v in vals):
        normalized = []
        for v in vals:
            if isinstance(v, datetime):
                normalized.append(v.date().isoformat())
            elif isinstance(v, date):
                normalized.append(v.isoformat())
            else:
                normalized.append(v)
        print(normalized)

print('TAB5_CONTENT_FRAMEWORK')
ws5 = wb['5. Content Framework']
for i, r in enumerate(ws5.iter_rows(values_only=True), 1):
    vals = list(r)
    if any(v is not None for v in vals):
        print(i, vals)
